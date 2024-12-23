import openpyxl
import os
import warnings
import math
import logging

# Suppress UserWarning for missing default style
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Setup logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# File handler for logging to a file
file_handler = logging.FileHandler('script_output.log')
file_handler.setLevel(logging.INFO)

# Stream handler for logging to terminal
stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.INFO)

# Formatter to format log messages
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
stream_handler.setFormatter(formatter)

# Add handlers to logger
logger.addHandler(file_handler)
logger.addHandler(stream_handler)

# Constants
NET_COST_STAR3 = 0  # maximum 3 star price difference
RATIO_COST_STAR3 = 0.05  # maximum 3 star price ratio
NET_COST_STAR2 = 0
RATIO_COST_STAR2 = 0.03
NET_COST_STAR1 = 0
RATIO_COST_STAR1 = 0.02
KEYWORDS = [""]
SPLIT = 7  # Number of split files

x1star = 0
x2star = 0
x3star = 0

# Directories
OUTPUT_DIR = 'avantajliurun'
SPLIT_DIR = os.path.join(OUTPUT_DIR, 'splited')
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)
if not os.path.exists(SPLIT_DIR):
    os.makedirs(SPLIT_DIR)

# Scan for Excel files
XLSX_FILES = [file for file in os.listdir("./")
              if not file.startswith("seller") and not file.startswith("\u00dcr\u00fcnleriniz") and file.endswith('.xlsx')]

logger.info(f"Found {len(XLSX_FILES)} files to process: {XLSX_FILES}")

for xlsx_file in XLSX_FILES:
    logger.info(f"Processing file: {xlsx_file}")

    # Load workbook and active sheet
    xlsx_path = os.path.join("./", xlsx_file)
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    # Get headers
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}

    # Create a new workbook for updated rows
    trimmed_workbook = openpyxl.Workbook()
    trimmed_sheet = trimmed_workbook.active
    trimmed_sheet.append([cell.value for cell in sheet[1]])  # Copy headers

    trimmed_rows = []  # Store rows to split later

    for row in sheet.iter_rows(min_row=2):
        try:
            # Get column values
            value1 = float(
                str(row[headers['TRENDYOL SATI\u015e F\u0130YATI'] - 1].value).replace(',', '.'))
            star1 = float(
                str(row[headers['1 YILDIZ \u00dcST F\u0130YAT'] - 1].value).replace(',', '.'))
            star2 = float(
                str(row[headers['2 YILDIZ \u00dcST F\u0130YAT'] - 1].value).replace(',', '.'))
            star3 = float(
                str(row[headers['3 YILDIZ \u00dcST F\u0130YAT'] - 1].value).replace(',', '.'))
            keyword = str(row[headers['MARKA'] - 1].value)

            # Determine price changes
            new_price = None
            if value1 - star3 < NET_COST_STAR3 or (value1 - star3) / value1 < RATIO_COST_STAR3:
                new_price = format(star3 - 0.01, '.2f').replace('.', ',')
                x3star += 1
            elif value1 - star2 < NET_COST_STAR2 or (value1 - star2) / value1 < RATIO_COST_STAR2:
                new_price = format(star2 - 0.01, '.2f').replace('.', ',')
                x2star += 1
            elif value1 - star1 < NET_COST_STAR1 or (value1 - star1) / value1 < RATIO_COST_STAR1:
                new_price = format(star1 - 0.01, '.2f').replace('.', ',')
                x1star += 1
            elif keyword in KEYWORDS:
                new_price = format(star1 - 0.01, '.2f').replace('.', ',')

            # Update rows with changes and mark for saving
            if new_price and row[headers['YEN\u0130 TSF (F\u0130YAT G\u00dcNCELLE)'] - 1].value != new_price:
                row[headers['YEN\u0130 TSF (F\u0130YAT G\u00dcNCELLE)'] -
                    1].value = new_price
                row[headers['Tarife Sonuna Kadar Uygula'] - 1].value = "Evet"
                trimmed_sheet.append([cell.value for cell in row])
                trimmed_rows.append([cell.value for cell in row])

        except KeyError as e:
            logger.error(f"Missing expected column: {e} in row {row[0].row}")
        except Exception as e:
            logger.error(f"Error processing row {row[0].row}: {str(e)}")

    # Save updated workbook
    trimmed_output_path = os.path.join(OUTPUT_DIR, f"trimmed_{xlsx_file}")
    if trimmed_sheet.max_row > 1:  # Check if any rows are added
        trimmed_workbook.save(trimmed_output_path)
        logger.info(f"Trimmed version saved to {trimmed_output_path}")
    else:
        logger.info(f"No changes detected in {
                    xlsx_file}, no trimmed file created.")

    # Split trimmed rows into 7 different files
    if trimmed_rows and SPLIT > 0:
        total_rows = len(trimmed_rows)
        rows_per_split = math.ceil(total_rows / SPLIT)

        for i in range(SPLIT):
            split_workbook = openpyxl.Workbook()
            split_sheet = split_workbook.active
            split_sheet.append(
                [cell.value for cell in sheet[1]])  # Add headers

            # Calculate start and end indices for this split
            start_idx = i * rows_per_split
            end_idx = min(start_idx + rows_per_split, total_rows)

            # Add rows to this split file one by one
            for row in trimmed_rows[start_idx:end_idx]:
                split_sheet.append(row)

            split_file_name = f"split_{xlsx_file.replace('.xlsx', '')}_{
                i + 1}.xlsx"
            split_output_path = os.path.join(SPLIT_DIR, split_file_name)
            split_workbook.save(split_output_path)
            logger.info(f"Split file saved to {split_output_path}")

    # Save the original workbook with changes
    output_file_path = os.path.join(OUTPUT_DIR, f"output_{xlsx_file}")
    workbook.save(output_file_path)
    logger.info(f"Original with updates saved to {output_file_path}")

# Summary of changes
logger.info(f"Changes Summary: New x1 Star Prices: {
            x1star}, New x2 Star Prices: {x2star}, New x3 Star Prices: {x3star}")
